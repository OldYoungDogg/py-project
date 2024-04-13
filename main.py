import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QTableWidget, QTableWidgetItem, QDialog, QLabel, QLineEdit, QMessageBox, QFormLayout, QAction, QFileDialog
from PyQt5.QtGui import QIcon, QFont
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import psycopg2
from cryptography.fernet import Fernet
from PyQt5.QtCore import pyqtSignal

# Версия программы
VERSION = "1.2"

class ComputerApp(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Учет компьютеров")
        self.setGeometry(100, 100, 800, 400)
        self.setWindowIcon(QIcon('icons.ico'))  
        self.init_ui()
        self.load_data() #загружаем данные при запуске программы
    def read_connection_info(self):
        with open("connection_info.txt", "rb") as f:
            encrypted_data = f.read()
        key = b'T7wknL4ZuLFpDlwqLR556kWdmwVWf0g5ZPL3HAB5iyk='  
        f = Fernet(key)
        decrypted_data = f.decrypt(encrypted_data)
        self.connection_info = eval(decrypted_data.decode())  
        return self.connection_info

    def init_ui(self):
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)

        self.layout = QVBoxLayout()
        self.central_widget.setLayout(self.layout)

        self.table = QTableWidget()
        self.layout.addWidget(self.table)

        self.table.setColumnCount(12)
        
        headers = [
            "Серийный номер компьютера",
            "Номер комнаты",
            "Тип устройства",
            "Модель устройства",
            "Наименование устройства",
            "Серийный номер устройства",
            "Процессор",
            "Серийный номер процессора",
            "Видеокарта",
            "Серийный номер видеокарты",
            "Диск",
            "Блок питания"
        ]
        self.table.setHorizontalHeaderLabels(headers)

        for i in range(self.table.columnCount()):
            header = self.table.horizontalHeaderItem(i)
            header.setText(header.text() if len(header.text()) <= 15 else "\n".join(header.text().split()))

        header_font = self.table.horizontalHeader().font()
        header_font.setPointSize(12)  
        self.table.horizontalHeader().setFont(header_font)  
        
        font = QFont()
        font.setPointSize(10)  
        self.table.setFont(font)  

        self.load_data()

        self.btn_add = QPushButton("Добавить запись")
        self.btn_delete = QPushButton("Удалить запись")
        self.btn_edit = QPushButton("Редактировать запись")

        font = QFont()
        font.setPointSize(14)  
        self.btn_add.setFont(font)
        self.btn_delete.setFont(font)
        self.btn_edit.setFont(font)

        self.btn_layout = QHBoxLayout()
        self.btn_layout.addWidget(self.btn_add)
        self.btn_layout.addWidget(self.btn_delete)
        self.btn_layout.addWidget(self.btn_edit)

        self.layout.addLayout(self.btn_layout)
        self.btn_add.clicked.connect(self.show_add_record_dialog)
        self.btn_delete.clicked.connect(self.delete_record)
        self.btn_edit.clicked.connect(self.show_edit_record_dialog)

        menubar = self.menuBar()
        import_menu = menubar.addMenu('Экспорт')
        version_menu = menubar.addMenu('Версия')
        refresh_menu = menubar.addMenu('Обновить')

        import_action = QAction('Экспорт', self)
        import_action.triggered.connect(self.export_data)
        import_menu.addAction(import_action)

        version_action = QAction('Версия', self)
        version_action.triggered.connect(self.show_version_info)
        version_menu.addAction(version_action)

        refresh_action = QAction('Обновить', self)
        refresh_action.triggered.connect(self.load_data)
        refresh_menu.addAction(refresh_action)


    def load_data(self):
        self.connection_info = self.read_connection_info()

        try:
            conn = psycopg2.connect(
                dbname=self.connection_info["dbname"],
                user=self.connection_info["user"],
                password=self.connection_info["password"],
                host=self.connection_info["host"],
                port=self.connection_info["port"]
            )

            cur = conn.cursor()

            cur.execute("""
            SELECT
                c.serial_num_pc AS "Серийный номер пк",
                c.room_num AS "Номер комнаты",
                d.device_type AS "Тип устройства",
                d.device_model AS "Модель устройства",
                d.device_name AS "Наименование устройства",
                d.device_serial AS "Серийный номер устройства",
                co.cpu AS "Процессор",
                co.serial_num_cpu AS "Серийный номер процессора",
                co.gpu AS "Видеокарта",
                co.serial_num_gpu AS "Серийный номер видеокарты",
                co.storage AS "Дисковое пространство",
                co.power_block AS "Блок питания" 
            FROM Computers c
            LEFT JOIN Devices d ON c.serial_num_pc = d.serial_num_pc
            LEFT JOIN Components co ON c.serial_num_pc = co.serial_num_pc;
            """)
            data = cur.fetchall()
            if data:
                self.table.setColumnCount(len(data[0]))
                self.table.setRowCount(len(data))
                for i, row in enumerate(data):
                    for j, value in enumerate(row):
                        item = QTableWidgetItem(str(value))
                        self.table.setItem(i, j, item)

            cur.close()
            conn.close()
        except psycopg2.Error as e:
            print("Error:", e)
            QMessageBox.critical(None, "Error", f"An error occurred: {e}")

    def export_data(self):
        file_path, _ = QFileDialog.getSaveFileName(self, 'Сохранить как', '', 'Excel Files (*.xlsx);;All Files (*)')
        if file_path:
            wb = Workbook()
            ws = wb.active

            headers = [self.table.horizontalHeaderItem(col).text() for col in range(self.table.columnCount())]

            for col, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col, value=header)
                ws.column_dimensions[get_column_letter(col)].width = len(header) + 2
                ws.cell(row=1, column=col).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

            for row in range(self.table.rowCount()):
                row_data = []
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    if item:
                        row_data.append(item.text())
                    else:
                        row_data.append("")
                ws.append(row_data)

            wb.save(file_path)

    def show_version_info(self):
        QMessageBox.information(self, "Версия программы", f" Версия программы: {VERSION}")

    def show_add_record_dialog(self):
        dialog = AddRecordDialog()
        if dialog.exec_():
            connection_info = self.read_connection_info()
            if connection_info:
                data = dialog.get_data()
                if data:
                    self.add_computer(connection_info, *data)
                self.load_data()
            else:
                QMessageBox.warning(self, "Ошибка", "Ошибка чтения информации о подключении.")

    def delete_record(self):
        selected_row = self.table.currentRow()
        if selected_row >= 0:
            items = []
            for i in range(self.table.columnCount()):
                items.append(self.table.item(selected_row, i).text())
            self.delete_computer(self.connection_info, items[0])
            self.table.setRowCount(0)
            self.load_data()
        else:
            QMessageBox.warning(self, "Ошибка", "Пожалуйста, выберите запись для удаления.")

    def delete_computer(self, connection_info, serial_num_pc):
        try:
            self.connection_info = self.read_connection_info()
            conn = psycopg2.connect(
                dbname=connection_info["dbname"],
                user=connection_info["user"],
                password=connection_info["password"],
                host=connection_info["host"],
                port=connection_info["port"]
            )

            cur = conn.cursor()

            cur.execute("DELETE FROM Computers WHERE serial_num_pc = %s", (serial_num_pc,))
            cur.execute("DELETE FROM Devices WHERE serial_num_pc = %s", (serial_num_pc,))
            cur.execute("DELETE FROM Components WHERE serial_num_pc = %s", (serial_num_pc,))
            cur.execute("DELETE FROM Rooms WHERE serial_num_pc = %s", (serial_num_pc,))
            cur.execute("DELETE FROM Components_Computers WHERE serial_num_pc = %s", (serial_num_pc,))

            conn.commit()
            cur.close()
            conn.close()

        except psycopg2.Error as e:
            conn.rollback()
            print("Error:", e)
            QMessageBox.critical(None, "Error", f"An error occurred: {e}")



    def show_edit_record_dialog(self):
        selected_row = self.table.currentRow()
        if selected_row != -1:
            items = []
            for i in range(self.table.columnCount()):
                items.append(self.table.item(selected_row, i).text())
            dialog = EditRecordDialog(self,items)
            dialog.set_update_function(self.update_data)
            # Создаем кнопку "Устройства" и подключаем к ней обработчик
            #devices_button = QPushButton("Устройства")
            #devices_button.clicked.connect(dialog.open_devices_dialog)
            #dialog.layout.addWidget(devices_button)
            
            dialog.exec_()
        else:
            QMessageBox.warning(self, "Ошибка", "Пожалуйста, выберите запись для редактирования.")

    def update_data(self, data):
        serial_num_pc, room_num, cpu, serial_num_cpu, gpu, serial_num_gpu, storage, power_block = data #device_type, device_model, device_name, device_serial, cpu
        print("Updating data:", data)  # Отладочное сообщение
        self.connection_info = self.read_connection_info()
        try:
            conn = psycopg2.connect(
                dbname=self.connection_info["dbname"],
                user=self.connection_info["user"],
                password=self.connection_info["password"],
                host=self.connection_info["host"],
                port=self.connection_info["port"]
            )

            cur = conn.cursor()

            #cur.execute("UPDATE Devices SET device_type = %s, device_model = %s, device_name = %s, device_serial = %s WHERE serial_num_pc = %s", (device_type, device_model, device_name, device_serial, serial_num_pc))
            cur.execute("UPDATE Components SET cpu = %s, serial_num_cpu = %s, gpu = %s, serial_num_gpu = %s, storage = %s, power_block = %s WHERE serial_num_pc = %s", (cpu, serial_num_cpu, gpu, serial_num_gpu, storage, power_block, serial_num_pc))
            cur.execute("UPDATE Computers SET room_num = %s WHERE serial_num_pc = %s", (room_num, serial_num_pc))

            conn.commit()
            cur.close()
            conn.close()

            self.load_data()
        except psycopg2.Error as e:
            print("Error:", e)
            QMessageBox.critical(None, "Error", f"An error occurred: {e}")

    def add_computer(self, connection_info, serial_num_pc, room_num, device_type, device_model, device_name, device_serial, cpu, serial_num_cpu, gpu, serial_num_gpu, storage, power_block):
        try:
            conn = psycopg2.connect(
                dbname=connection_info["dbname"],
                user=connection_info["user"],
                password=connection_info["password"],
                host=connection_info["host"],
                port=connection_info["port"]
            )

            cur = conn.cursor()

            cur.execute("INSERT INTO Computers (serial_num_pc, room_num) VALUES (%s, %s)", (serial_num_pc, room_num))
            cur.execute("INSERT INTO Rooms (room_num, serial_num_pc, pc_id) VALUES (%s, %s, (SELECT pc_id FROM Computers WHERE serial_num_pc = %s))", (room_num, serial_num_pc, serial_num_pc))
            cur.execute("INSERT INTO Devices (device_type, device_serial, device_model, device_name, serial_num_pc) VALUES (%s, %s, %s, %s, %s)", (device_type, device_serial, device_model, device_name, serial_num_pc))
            cur.execute("INSERT INTO Components (cpu, gpu, storage, power_block, serial_num_pc, serial_num_cpu, serial_num_gpu) VALUES (%s, %s, %s, %s, %s, %s, %s)", (cpu, gpu, storage, power_block, serial_num_pc, serial_num_cpu, serial_num_gpu))
            cur.execute("INSERT INTO Components_Computers (component_id, serial_num_pc) VALUES ((SELECT component_id FROM Components WHERE serial_num_pc = %s), %s)", (serial_num_pc, serial_num_pc))

            conn.commit()
            cur.close()
            conn.close()
        except psycopg2.Error as e:
            print("Error:", e)
            QMessageBox.critical(None, "Error", f"An error occurred: {e}")

class AddRecordDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Добавление записи")
        self.setWindowIcon(QIcon("icons.ico"))

        self.layout = QVBoxLayout()
        self.form_layout = QFormLayout()

        self.serial_num_pc_input = QLineEdit()
        self.room_num_input = QLineEdit()
        self.device_type_input = QLineEdit()
        self.device_model_input = QLineEdit()
        self.device_name_input = QLineEdit()
        self.device_serial_input = QLineEdit()
        self.cpu_input = QLineEdit()
        self.serial_num_cpu_input = QLineEdit()
        self.gpu_input = QLineEdit()
        self.serial_num_gpu_input = QLineEdit()
        self.storage_input = QLineEdit()
        self.power_block_input = QLineEdit()

        self.form_layout.addRow("Серийный номер компьютера:", self.serial_num_pc_input)
        self.form_layout.addRow("Номер комнаты:", self.room_num_input)
        self.form_layout.addRow("Тип устройства:", self.device_type_input)
        self.form_layout.addRow("Модель устройства:", self.device_model_input)
        self.form_layout.addRow("Наименование устройства:", self.device_name_input)
        self.form_layout.addRow("Серийный номер устройства:", self.device_serial_input)
        self.form_layout.addRow("Процессор:", self.cpu_input)
        self.form_layout.addRow("Серийный номер процессора:", self.serial_num_cpu_input)
        self.form_layout.addRow("Видеокарта:", self.gpu_input)
        self.form_layout.addRow("Серийный номер видеокарты:", self.serial_num_gpu_input)
        self.form_layout.addRow("Дисковое пространство:", self.storage_input)
        self.form_layout.addRow("Блок питания:", self.power_block_input)

        self.layout.addLayout(self.form_layout)

        self.add_button = QPushButton("Добавить")
        self.add_button.clicked.connect(self.accept)

        self.layout.addWidget(self.add_button)
        self.setLayout(self.layout)

    def get_data(self):
        serial_num_pc = self.serial_num_pc_input.text()
        room_num = self.room_num_input.text()
        device_type = self.device_type_input.text()
        device_model = self.device_model_input.text()
        device_name = self.device_name_input.text()
        device_serial = self.device_serial_input.text()
        cpu = self.cpu_input.text()
        serial_num_cpu = self.serial_num_cpu_input.text()
        gpu = self.gpu_input.text()
        serial_num_gpu = self.serial_num_gpu_input.text()
        storage = self.storage_input.text()
        power_block = self.power_block_input.text()

        return serial_num_pc, room_num, device_type, device_model, device_name, device_serial, cpu, serial_num_cpu, gpu, serial_num_gpu, storage, power_block


class EditRecordDialog(QDialog):

    def __init__(self, parent, data):
        super().__init__(parent)
        self.parent = parent
        self.setWindowTitle("Редактирование записи")
        
        self.setWindowIcon(QIcon("icons.ico"))

        self.layout = QVBoxLayout()

        self.form_layout = QFormLayout()

        self.serial_num_pc_label = QLabel("Серийный номер компьютера:")
        self.serial_num_pc_input = QLineEdit(data[0])

        self.room_num_label = QLabel("Номер комнаты:")
        self.room_num_input = QLineEdit(data[1])

        #self.device_serial_label = QLabel("Серийный номер устройства:")
        #self.device_serial_input = QLineEdit(data[5])

        #self.device_type_label = QLabel("Тип устройства:")
        #self.device_type_input = QLineEdit(data[2])

        #self.device_model_label = QLabel("Модель устройства:")
        #self.device_model_input = QLineEdit(data[3])

        #self.device_name_label = QLabel("Наименование устройства:")
        #self.device_name_input = QLineEdit(data[4])

        self.cpu_label = QLabel("Процессор:")
        self.cpu_input = QLineEdit(data[6])

        self.serial_num_cpu_label = QLabel("Серийный номер процессора:")
        self.serial_num_cpu_input = QLineEdit(data[7])

        self.gpu_label = QLabel("Видеокарта:")
        self.gpu_input = QLineEdit(data[8])

        self.serial_num_gpu_label = QLabel("Серийный номер видеокарты:")
        self.serial_num_gpu_input = QLineEdit(data[9])

        self.storage_label = QLabel("Дисковое пространство:")
        self.storage_input = QLineEdit(data[10])

        self.power_block_label = QLabel("Блок питания:")
        self.power_block_input = QLineEdit(data[11])

        # Добавляем поля ввода в form layout
        self.form_layout.addRow(self.serial_num_pc_label, self.serial_num_pc_input)
        self.form_layout.addRow(self.room_num_label, self.room_num_input)
        #self.form_layout.addRow(self.device_serial_label, self.device_serial_input)
        #self.form_layout.addRow(self.device_type_label, self.device_type_input)
        #self.form_layout.addRow(self.device_model_label, self.device_model_input)
        #self.form_layout.addRow(self.device_name_label, self.device_name_input)
        self.form_layout.addRow(self.cpu_label, self.cpu_input)
        self.form_layout.addRow(self.serial_num_cpu_label, self.serial_num_cpu_input)
        self.form_layout.addRow(self.gpu_label, self.gpu_input)
        self.form_layout.addRow(self.serial_num_gpu_label, self.serial_num_gpu_input)
        self.form_layout.addRow(self.storage_label, self.storage_input)
        self.form_layout.addRow(self.power_block_label, self.power_block_input)

        self.layout.addLayout(self.form_layout)

        self.save_button = QPushButton("Сохранить")
        self.save_button.clicked.connect(self.save_data)

        # Добавляем кнопку "Сохранить" в layout
        self.layout.addWidget(self.save_button)
        # Добавляем кнопку "Устройства" На форму редактирования Записи
        self.devices_button = QPushButton("Устройства")
        self.devices_button.clicked.connect(self.open_devices_dialog)
        self.layout.addWidget(self.devices_button)

        self.setLayout(self.layout)

        self.data = data
        self.update_function = None

#Устройства
    def open_devices_dialog(self):
        devices_dialog = DevicesDialog(self, self.data[0])

        devices_dialog.exec_()





    def set_update_function(self, update_function):
        self.update_function = update_function

    def save_data(self):
        if self.update_function:
            new_data = (
                self.serial_num_pc_input.text(),
                self.room_num_input.text(),
                #self.device_type_input.text(),
                #self.device_model_input.text(),
                #self.device_name_input.text(),
                #self.device_serial_input.text(),
                self.cpu_input.text(),
                self.serial_num_cpu_input.text(),
                self.gpu_input.text(),
                self.serial_num_gpu_input.text(),
                self.storage_input.text(),
                self.power_block_input.text()
            )
            
            self.update_function(new_data)
            self.accept()

    def get_data(self):
        serial_num_pc = self.serial_num_pc_input.text()
        room_num = self.room_num_input.text()
        #device_type = self.device_type_input.text()
        #device_model = self.device_model_input.text()
        #device_name = self.device_name_input.text()
        #device_serial = self.device_serial_input.text()
        cpu = self.cpu_input.text()
        serial_num_cpu = self.serial_num_cpu_input.text()
        gpu = self.gpu_input.text()
        serial_num_gpu = self.serial_num_gpu_input.text()
        storage = self.storage_input.text()
        power_block = self.power_block_input.text()
        return serial_num_pc, room_num, cpu, serial_num_cpu, gpu, serial_num_gpu, storage, power_block #device_type, device_model, device_name, device_serial, cpu

#Класс Для отображения Формы Устройства. Удаление и добавление записей формы Устройства (devices)
class DevicesDialog(QDialog):
    device_deleted = pyqtSignal()
    def __init__(self, parent, serial_num_pc):
        super().__init__(parent)
        self.devices = []

        self.serial_num_pc = serial_num_pc
        self.setWindowTitle("Устройства компьютера")
        self.devices_table = QTableWidget()
        self.add_button = QPushButton("Добавить")
        self.edit_button = QPushButton("Редактировать")
        self.delete_button = QPushButton("Удалить")
        self.add_button.clicked.connect(self.add_device)
        self.edit_button.clicked.connect(self.edit_device)
        self.delete_button.clicked.connect(self.delete_device)
        

        # Создаем таблицу для отображения устройств
        self.devices_table = QTableWidget()
        self.devices_table.setColumnCount(5)  # Устанавливаем количество столбцов
        self.devices_table.setHorizontalHeaderLabels(["Серийный номер компьютера", "Тип устройства", "Наименование устройства", "Модель устройства", "Серийный номер устройства"])
        self.layout = QHBoxLayout()
        self.layout = QVBoxLayout()
        self.layout.addWidget(self.devices_table)
        self.layout.addWidget(self.add_button)
        self.layout.addWidget(self.edit_button)
        self.layout.addWidget(self.delete_button)
        self.setLayout(self.layout)

        # Выполняем SQL-запрос для получения устройств по серийному номеру компьютера
        self.load_devices()

    def read_connection_info(self):
        with open("connection_info.txt", "rb") as f:
            encrypted_data = f.read()
        key = b'T7wknL4ZuLFpDlwqLR556kWdmwVWf0g5ZPL3HAB5iyk='  
        f = Fernet(key)
        decrypted_data = f.decrypt(encrypted_data)
        self.connection_info = eval(decrypted_data.decode())  
        return self.connection_info



    def load_devices(self):
        
        self.connection_info = self.read_connection_info()
        try:
            conn = psycopg2.connect(
                dbname=self.connection_info["dbname"],
                user=self.connection_info["user"],
                password=self.connection_info["password"],
                host=self.connection_info["host"],
                port=self.connection_info["port"]
            )

            cur = conn.cursor()

        # Выполнение SQL-запроса для получения устройств по серийному номеру компьютера
            cur.execute("""
                SELECT c.serial_num_pc AS "Серийный номер компьютера",
                        d.device_type AS "Тип устройства",
                        d.device_name AS "Наименование устройства",
                        d.device_model AS "Модель устройства",
                        d.device_serial AS "Серийный номер устройства"
                FROM Computers c
                LEFT JOIN Devices d ON c.serial_num_pc = d.serial_num_pc
                WHERE c.serial_num_pc = %s
            """, (self.serial_num_pc,))
            result = cur.fetchall()
            #self.devices = []
            self.devices_table.setRowCount(0)
            if result:
                self.devices_table.setRowCount(len(result))
                for row_idx, row in enumerate(result):
                    for col_idx, data in enumerate(row):
                        item = QTableWidgetItem(str(data))
                        self.devices_table.setItem(row_idx, col_idx, item)
        except Exception as e:
            print("Error:", e)

    def add_device(self):
    # Создаем диалоговое окно для добавления устройства
        add_device_dialog = AddDeviceDialog(self)
        if add_device_dialog.exec_():
        # Если пользователь нажал "ОК", получаем данные из диалогового окна
            device_data = add_device_dialog.get_device_data()

        # Проверяем, что данные устройства не пустые
            if device_data:
            # Добавляем новое устройство в базу данных
                conn = None
                cur = None
                try:
                    conn = psycopg2.connect(
                        dbname=self.connection_info["dbname"],
                        user=self.connection_info["user"],
                        password=self.connection_info["password"],
                        host=self.connection_info["host"],
                        port=self.connection_info["port"]
                    )
                    cur = conn.cursor()

                # Выполняем SQL-запрос для добавления нового устройства
                    cur.execute("""
                        INSERT INTO Devices (serial_num_pc, device_type, device_name, device_model, device_serial)
                        VALUES (%s, %s, %s, %s, %s)
                    """, (self.serial_num_pc, *device_data))
                
                    conn.commit()
                # После успешного добавления обновляем отображение устройств
                    self.load_devices()
                    QMessageBox.information(self, "Успех", "Устройство успешно добавлено.")
                except Exception as e:
                    print("Error:", e)
                    QMessageBox.critical(None, "Error", f"An error occurred: {e}")
                finally:
                    if cur:
                        cur.close()
                    if conn:
                        conn.close()
        
        
    def delete_device(self):
        selected_row = self.devices_table.currentRow()
        if selected_row != -1:
            # Получаем серийный номер устройства из второго столбца выбранной строки
            serial_num_device = self.devices_table.item(selected_row, 4).text()  # Изменил индекс столбца на 4
            print("Deleting device with serial number:", serial_num_device)
            # Получаем идентификатор устройства
            device_id = self.get_device_id(serial_num_device) 
            print("Delete device with device_id:", device_id)
            if device_id is not None:
                # Удаляем устройство из базы данных
                conn = None
                cur = None
                self.connection_info = self.read_connection_info()
                try:
                    conn = psycopg2.connect(
                        dbname=self.connection_info["dbname"],
                        user=self.connection_info["user"],
                        password=self.connection_info["password"],
                        host=self.connection_info["host"],
                        port=self.connection_info["port"]
                    )
                    cur = conn.cursor()
            
                    # Выполняем SQL-запрос для удаления устройства по его идентификатору
                    cur.execute("DELETE FROM Devices WHERE device_id = %s", (device_id,))
                    conn.commit()
                    # После успешного удаления отправляем сигнал
                    self.device_deleted.emit()

                    # После успешного удаления обновляем отображение устройств
                    self.load_devices()
                    
                    QMessageBox.information(self, "Успех", "Устройство успешно удалено.")
                except Exception as e:
                    print("Error:", e)
                    QMessageBox.critical(None, "Error", f"An error occurred: {e}")
                finally:
                    if cur:
                        cur.close()
                    if conn:
                        conn.close()
            else:
                QMessageBox.warning(self, "Ошибка", "Не удалось получить идентификатор устройства.")
        else:
            QMessageBox.warning(self, "Ошибка", "Пожалуйста, выберите устройство для удаления.")




    def get_device_id(self, serial_num_device):
        conn = None
        cur = None
        device_id = None
        try:
            self.connection_info = self.read_connection_info()
            conn = psycopg2.connect(
                dbname=self.connection_info["dbname"],
                user=self.connection_info["user"],
                password=self.connection_info["password"],
                host=self.connection_info["host"],
                port=self.connection_info["port"]
            )
            cur = conn.cursor()
            cur.execute("SELECT device_id FROM Devices WHERE device_serial = %s", (serial_num_device,))
            device_id = cur.fetchone()[0]  # Получаем первый элемент кортежа
        except psycopg2.Error as e:
            print("Error:", e)
        finally:
            if cur:
                cur.close()
            if conn:
                conn.close()
        return device_id

    def edit_device(self):
        selected_row = self.devices_table.currentRow()
        if selected_row != -1:
            device_id = self.devices_table.item(selected_row, 0).text()
            device_type = self.devices_table.item(selected_row, 1).text()
            device_name = self.devices_table.item(selected_row, 2).text()
            device_model = self.devices_table.item(selected_row, 3).text()
            device_serial = self.devices_table.item(selected_row, 4).text()

            edit_dialog = EditDeviceDialog(device_id,[device_type, device_model, device_name, device_serial], self)
            if edit_dialog.exec_():
                self.load_devices()
        else:
            QMessageBox.warning(self, "Ошибка", "Пожалуйста, выберите устройство для редактирования.")
            
#Класс для создания окна добавления Устройства к компьютера из Формы Редактирования записи
class AddDeviceDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Добавить устройство")
        
        layout = QVBoxLayout()
        
        self.device_type_edit = QLineEdit()
        self.device_model_edit = QLineEdit()
        self.device_name_edit = QLineEdit()
        self.device_serial_edit = QLineEdit()
        
        layout.addWidget(QLabel("Тип устройства:"))
        layout.addWidget(self.device_type_edit)
        layout.addWidget(QLabel("Модель устройства:"))
        layout.addWidget(self.device_model_edit)
        layout.addWidget(QLabel("Наименование устройства:"))
        layout.addWidget(self.device_name_edit)
        layout.addWidget(QLabel("Серийный номер устройства:"))
        layout.addWidget(self.device_serial_edit)
        
        self.ok_button = QPushButton("OK")
        self.ok_button.clicked.connect(self.accept)
        layout.addWidget(self.ok_button)
        
        self.setLayout(layout)
        
    def get_device_data(self):
        device_type = self.device_type_edit.text()
        device_model = self.device_model_edit.text()
        device_name = self.device_name_edit.text()
        device_serial = self.device_serial_edit.text()
        
        # Проверяем, что все поля заполнены
        if device_type and device_model and device_name and device_serial:
            return device_type, device_model, device_name, device_serial
        else:
            return None

class EditDeviceDialog(QDialog):
    def __init__(self, device_id, data, parent=None):
        super().__init__(parent)
        self.device_id = device_id
        #self.serial_num_pc = serial_num_pc
        self.setWindowTitle("Редактирование устройства")
        
        layout = QVBoxLayout()
        
        self.device_type_edit = QLineEdit(data[0])
        self.device_model_edit = QLineEdit(data[1])
        self.device_name_edit = QLineEdit(data[2])
        self.device_serial_edit = QLineEdit(data[3])
        
        form_layout = QFormLayout()
        form_layout.addRow("Тип устройства:", self.device_type_edit)
        form_layout.addRow("Модель устройства:", self.device_model_edit)
        form_layout.addRow("Наименование устройства:", self.device_name_edit)
        form_layout.addRow("Серийный номер устройства:", self.device_serial_edit)
        
        layout.addLayout(form_layout)
        
        self.ok_button = QPushButton("OK")
        self.ok_button.clicked.connect(self.save_data)  # Перенесли подключение к функции сохранения данных
        layout.addWidget(self.ok_button)
        
        self.setLayout(layout)
    
    def read_connection_info(self):
        with open("connection_info.txt", "rb") as f:
            encrypted_data = f.read()
        key = b'T7wknL4ZuLFpDlwqLR556kWdmwVWf0g5ZPL3HAB5iyk='  
        f = Fernet(key)
        decrypted_data = f.decrypt(encrypted_data)
        self.connection_info = eval(decrypted_data.decode())  
        return self.connection_info


    def save_data(self):
        # Получаем измененные данные из полей ввода
        device_type = self.device_type_edit.text()
        device_model = self.device_model_edit.text()
        device_name = self.device_name_edit.text()
        device_serial = self.device_serial_edit.text()
        # Получаем device_id из базы данных
        device_id = self.get_device_id()  # Используем метод get_device_id без аргументов
        # Здесь нужно выполнить SQL-запрос для обновления данных в базе данных
        if device_id is not None:
            self.connection_info = self.read_connection_info()
            try:
                conn = psycopg2.connect(
                    dbname=self.connection_info["dbname"],
                    user=self.connection_info["user"],
                    password=self.connection_info["password"],
                    host=self.connection_info["host"],
                    port=self.connection_info["port"]
                )

                cur = conn.cursor()
                cur.execute("UPDATE Devices SET device_type = %s, device_model = %s, device_name = %s, device_serial = %s WHERE device_id = %s", (device_type, device_model, device_name, device_serial, device_id))
                print("в апдете получаю нужный ид:",device_id)
                conn.commit()
                cur.close()
                conn.close()

                #self.load_data()
            except psycopg2.Error as e:
                print("Error:", e)
                QMessageBox.critical(None, "Error", f"An error occurred: {e}")

            # Закрываем диалоговое окно после сохранения данных
            self.accept()
        else:
            QMessageBox.warning(None,"Warning", "Unable to retrieve device_id.")

    
    def get_device_id(self):
        device_serial = self.device_serial_edit.text()

        try:
            self.connection_info = self.read_connection_info()
            conn = psycopg2.connect(
                dbname=self.connection_info["dbname"],
                user=self.connection_info["user"],
                password=self.connection_info["password"],
                host=self.connection_info["host"],
                port=self.connection_info["port"]
            )
            cur = conn.cursor()

            # Выполнение запроса к базе данных для получения device_id по серийному номеру устройства
            cur.execute("SELECT device_id FROM Devices WHERE device_serial = %s", (device_serial,))
            result = cur.fetchone()

            cur.close()
            conn.close()

            if result:
                return result[0]  # Возвращаем найденный device_id
            else:
                return None  # Если устройство с указанным серийным номером не найдено, возвращаем None
        except psycopg2.Error as e:
            print("Error:", e)
            QMessageBox.critical(None, "Error", f"An error occurred while retrieving device_id: {e}")
            return None  # В случае ошибки вернем None

        

app = QApplication(sys.argv)
window = ComputerApp()
window.show()
sys.exit(app.exec_())
